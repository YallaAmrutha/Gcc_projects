import openpyxl

def add_ToastMaster_data():
    ToastMasters = []

    while True:
        print("Enter ToastMaster details:")
        name = input("Name: ")
        ToastMaster_id = input("ToastMaster_id: ")
        position = input("Position:")
        phone_no = input("Phone No: ")

        ToastMaster = {
            "Name": name,
            "ToastMaster_id": ToastMaster_id,
            "position": position,
            "Phone No": phone_no,
        }
        ToastMasters.append(ToastMaster)

        more_data = input("Do you want to enter data for another ToastMaster Details? (yes/no): ")
        if more_data.lower() != "yes":
            break

    return ToastMasters

def create_excel(ToastMasters, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write headers
    headers = list(ToastMasters[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    # Write student data
    for row_idx, ToastMaster in enumerate(ToastMasters, start=2):
        for col_idx, value in enumerate(ToastMaster.values(), start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # Save the Excel file
    workbook.save(output_file)

if __name__ == "__main__":
    print("Enter student data. Type 'done' when you finish.")

    ToastMaster_data = add_ToastMaster_data()

    if ToastMaster_data:
        print("\nToastMaster data collected successfully.")
        output_file_name = input("Enter the output Excel file name: ")

        create_excel(ToastMaster_data, output_file_name)
        print(f"Data successfully converted and saved to {output_file_name}.")
    else:
        print("No ToastMaster data to convert.")